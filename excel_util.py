from openpyxl import load_workbook

def read_test_data(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))
    return data

def write_test_result(file_path, test_id, result):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == test_id:
            ws.cell(row, 7).value = result
            break
    wb.save(file_path)
